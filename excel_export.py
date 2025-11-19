"""
Modulo per l'esportazione dei dati in formato Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class ExcelExporter:
    """Classe per esportare i dati delle verifiche in Excel"""

    def __init__(self):
        """Inizializza l'esportatore"""
        pass

    def export_to_excel(self, data, filename=None):
        """
        Esporta i dati in un file Excel

        Args:
            data (list): Lista di dizionari con i dati delle verifiche
            filename (str): Nome del file (opzionale)

        Returns:
            str: Nome del file creato
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"vat_checks_{timestamp}.xlsx"

        # Assicurati che il filename abbia l'estensione .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        # Crea un nuovo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Verifiche P.IVA"

        # Definisci gli headers
        headers = [
            'ID',
            'Partita IVA',
            'Codice Paese',
            'Valida',
            'Nome/Ragione Sociale',
            'Indirizzo',
            'P.IVA Richiedente',
            'Data Verifica',
            'Errore'
        ]

        # Stile per gli headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Scrivi gli headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Scrivi i dati
        for row_num, record in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=record.get('id', ''))
            ws.cell(row=row_num, column=2, value=record.get('vat_number', ''))
            ws.cell(row=row_num, column=3, value=record.get('country_code', ''))

            # Colonna valida con formattazione
            valid_cell = ws.cell(row=row_num, column=4)
            if record.get('valid'):
                valid_cell.value = 'SI'
                valid_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                valid_cell.font = Font(color="006100")
            else:
                valid_cell.value = 'NO'
                valid_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                valid_cell.font = Font(color="9C0006")

            ws.cell(row=row_num, column=5, value=record.get('name', ''))
            ws.cell(row=row_num, column=6, value=record.get('address', ''))
            ws.cell(row=row_num, column=7, value=record.get('requester_vat', ''))
            ws.cell(row=row_num, column=8, value=record.get('request_date', ''))
            ws.cell(row=row_num, column=9, value=record.get('error', ''))

        # Regola la larghezza delle colonne
        column_widths = {
            'A': 8,   # ID
            'B': 18,  # Partita IVA
            'C': 14,  # Codice Paese
            'D': 10,  # Valida
            'E': 35,  # Nome
            'F': 50,  # Indirizzo
            'G': 18,  # P.IVA Richiedente
            'H': 20,  # Data
            'I': 30   # Errore
        }

        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width

        # Blocca la prima riga (headers)
        ws.freeze_panes = 'A2'

        # Salva il file
        wb.save(filename)

        return filename

    def export_summary(self, stats, filename=None):
        """
        Esporta un riepilogo statistico in Excel

        Args:
            stats (dict): Dizionario con le statistiche
            filename (str): Nome del file (opzionale)

        Returns:
            str: Nome del file creato
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"vat_summary_{timestamp}.xlsx"

        if not filename.endswith('.xlsx'):
            filename += '.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Riepilogo"

        # Titolo
        ws['A1'] = 'RIEPILOGO VERIFICHE PARTITE IVA'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Statistiche
        row = 3
        stats_data = [
            ('Totale verifiche:', stats.get('total_checks', 0)),
            ('Verifiche valide:', stats.get('valid_checks', 0)),
            ('Verifiche non valide:', stats.get('invalid_checks', 0)),
            ('Partite IVA uniche:', stats.get('unique_vats', 0))
        ]

        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Regola larghezza colonne
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        wb.save(filename)

        return filename
